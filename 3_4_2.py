import pandas as pd
from concurrent import futures
import os
import sys
from numpy import mean
import math
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
import numpy as np
import matplotlib.pyplot as plt
from jinja2 import Environment, FileSystemLoader
import pdfkit


def analysis_statistics(vacancy, year):
    st_df = pd.read_csv(f'csv_files\\part_{year}.csv')
    st_df.loc[:, 'salary'] = st_df.loc[:, ['salary_from', 'salary_to']].mean(axis=1)
    st_df_vacancy = st_df[st_df["name"].str.contains(vacancy)]

    salary_lvl_year = {year: []}
    count_vac_year = {year: 0}
    salary_lvl_year_prof = {year: []}
    count_vac_year_prof = {year: 0}

    salary_lvl_year[year] = int(st_df['salary'].mean())
    count_vac_year[year] = len(st_df)
    salary_lvl_year_prof[year] = int(st_df_vacancy['salary'].mean())
    count_vac_year_prof[year] = len(st_df_vacancy)

    return [salary_lvl_year, count_vac_year, salary_lvl_year_prof, count_vac_year_prof]


if __name__ == "__main__":
    class UserInput:
        """Класс для пользовательского ввода.

        Attributes:
            file_name (str): Название файла
            profession_name (str): Название професии
        """
        def __init__(self):
            """Инициализирует объекты UserInput, выполняет проверку на корректность введенные данные.
            """
            self.file_name = input('Введите название файла: ')
            self.profession_name = input('Введите название профессии: ')

            self.file_name = self.check_file_name(self.file_name)
            self.profession_name = self.check_profession_name(self.profession_name)

        @staticmethod
        def check_file_name(file_name):
            """Осуществляет проверку корретности названия файла.

            Returns:
                str: Название файла
            """
            if file_name == '' or '.' not in file_name:
                print('Некорректное название файла')
                sys.exit()
            return file_name

        @staticmethod
        def check_profession_name(profession_name):
            """Осуществляет проверку корретности названия профессии.

            Returns:
                str: Название профессии
            """
            if profession_name == '':
                print('Некорректное название профессии')
                sys.exit()
            return profession_name


    class GetCsvParts:
        def __init__(self, file_name):
            self.df = pd.read_csv(file_name)
            self.df['years'] = self.df['published_at'].apply(lambda s: s[:4])
            self.years = self.df['years'].unique()

            for year in self.years:
                data = self.df[self.df['years'] == year]
                data.iloc[:, :6].to_csv(rf'csv_files\part_{year}.csv', index=False)


    class Salary:
        def __init__(self, df):
            self.df = df
            self.df_date = pd.read_csv("currency_date.csv")
            self.df["salary"] = df.apply(lambda row: self.get_salary(row["salary_from"], row["salary_to"], row["salary_currency"], row["published_at"][:7].split("-")), axis=1)

        def get_salary(self, salary_from, salary_to, salary_currency, date):
            date = date[1] + "/" + date[0]

            salary_currency_coef = 0
            if salary_currency == "RUR":
                salary_currency_coef = 1
            elif salary_currency != "RUR" and (salary_currency == salary_currency) and salary_currency in ["BYN", "BYR", "USD", "EUR", "KZT", "UAH"]:
                salary_currency.replace("BYN", "BYR")
                df_date_row = self.df_date.loc[self.df_date["date"] == date]
                salary_currency_coef = df_date_row[salary_currency].values[0]

            if not (math.isnan(salary_from)) and math.isnan(salary_to):
                return salary_from * salary_currency_coef
            elif math.isnan(salary_from) and not (math.isnan(salary_to)):
                return salary_to * salary_currency_coef
            elif not (math.isnan(salary_from)) and not (math.isnan(salary_to)):
                return mean([salary_from, salary_to]) * salary_currency_coef


    class Report:
        """Класс для формирования графиков и отчетов.

        Attributes:
            salary_lvl_by_year (SalaryDict): Динамика уровня зарплат по годам
            count_vac_by_year (CountDict): Динамика количества вакансий по годам
            salary_lvl_by_year_for_prof (SalaryDict): Динамика уровня зарплат по годам для выбранной профессии
            count_vac_by_year_for_prof (CountDict): Динамика количества вакансий по годам для выбранной профессии
            salary_lvl_by_city (SalaryDict): Уровень зарплат по городам
            vacancy_rate_by_city (CountDict): Доля вакансий по городам
            prof (str): Название профессии
            wb (Workbook): Объект книги (контейнер для всех остальных частей XLSX-документа)
            sheet1 (Worksheet): Лист для статистики по годам
            sheet2 (Worksheet): Лист для статистики по городам
            fig (Figure): область Figure (внешний контейнер для графики matplotlib)
            ax1 (): Добавленная к Figure область Axes (график уровеня зарплат по годам)
            ax2 (): Добавленная к Figure область Axes (график количества вакансий по годам)
            ax3 (): Добавленная к Figure область Axes (график уровеня зарплат по городам)
            ax4 (): Добавленная к Figure область Axes (график количества вакансий по городам)
        """

        def __init__(self, data, prof):
            """Инициализирует объекты Report.

            Args:
                data (list): Статистика вакансий
                prof (str): Название профессии
            """
            self.salary_lvl_by_year = data[0]
            self.count_vac_by_year = data[1]
            self.salary_lvl_by_year_for_prof = data[2]
            self.count_vac_by_year_for_prof = data[3]
            self.salary_lvl_by_city = data[4]
            self.vacancy_rate_by_city = data[5]
            self.prof = prof

            self.wb = Workbook()
            self.sheet1 = self.wb.active
            self.sheet1.title = 'Статистика по годам'
            self.sheet2 = self.wb.create_sheet('Статистика по городам')

            self.fig = plt.figure()
            self.ax1 = self.fig.add_subplot(221)
            self.ax1.set_title('Уровень зарплат по годам')
            self.ax2 = self.fig.add_subplot(222)
            self.ax2.set_title('Количество вакансий по годам')
            self.ax3 = self.fig.add_subplot(223)
            self.ax3.set_title('Уровень зарплат по городам')
            self.ax4 = self.fig.add_subplot(224)
            self.ax4.set_title('Доля вакансий по городам')

        def generate_excel(self):
            """Осуществляет генерацию данных ecxel файла, содержащего две вкладки
            "Статистика по годам" и "Статистика по городам".

            Returns:
                : Данные для ecxel файл
            """
            names_sheet1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.prof}',
                            'Количество вакансий', f'Количество вакансий - {self.prof}']
            names_sheet2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']

            for i, name in enumerate(names_sheet1):
                self.sheet1.cell(row=1, column=(i + 1), value=name).font = Font(bold=True)
            for year, value in self.salary_lvl_by_year.items():
                self.sheet1.append([year, value, self.salary_lvl_by_year_for_prof[year], self.count_vac_by_year[year],
                                    self.count_vac_by_year_for_prof[year]])

            for i, name in enumerate(names_sheet2):
                self.sheet2.cell(row=1, column=(i + 1), value=name).font = Font(bold=True)
            for i in range(len(list(self.salary_lvl_by_city.keys()))):
                self.sheet2.append([list(self.salary_lvl_by_city.keys())[i], list(self.salary_lvl_by_city.values())[i],
                                    list(self.vacancy_rate_by_city.keys())[i],
                                    list(self.vacancy_rate_by_city.values())[i]])

            side = Side(border_style='thin', color='000000')
            self.set_border(self.sheet1, side)
            self.set_border(self.sheet2, side)
            self.sheet2.insert_cols(3)
            self.sheet2.column_dimensions['C'].width = 2

            self.column_width(self.sheet1)
            self.column_width(self.sheet2)

            for i in range(2, len(self.sheet2['E']) + 1):
                self.sheet2[f'E{i}'].number_format = FORMAT_PERCENTAGE_00

        @staticmethod
        def set_border(ws, side):
            """Добавляет границы для ячеек.
            """
            for cell in ws._cells.values():
                cell.border = Border(top=side, bottom=side, left=side, right=side)

        @staticmethod
        def column_width(ws):
            """Устанавливает ширину ячейки.
            """
            dims = {}
            for row in ws.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))
            for col, value in dims.items():
                ws.column_dimensions[col].width = value + 2

        def generate_image(self):
            """Осуществляет генерацию png файла с четырмя графиками (уровня зарплат/вакансий, по годам/городам).

            Returns:
                : png файл
            """
            width_12 = 0.4
            x_nums_1 = np.arange(len(self.salary_lvl_by_year.keys()))
            x_list1_1 = x_nums_1 - width_12 / 2
            x_list1_2 = x_nums_1 + width_12 / 2

            self.ax1.bar(x_list1_1, self.salary_lvl_by_year.values(), width_12, label='средняя з/п')
            self.ax1.bar(x_list1_2, self.salary_lvl_by_year_for_prof.values(), width_12, label=f'з/п {self.prof}')
            self.ax1.set_xticks(x_nums_1, self.salary_lvl_by_year.keys(), rotation='vertical')
            self.ax1.tick_params(axis='both', labelsize=8)
            self.ax1.legend(fontsize=8)
            self.ax1.grid(True, axis='y')

            x_nums_2 = np.arange(len(self.count_vac_by_year.keys()))
            x_list2_1 = x_nums_2 - width_12 / 2
            x_list2_2 = x_nums_2 + width_12 / 2

            self.ax2.bar(x_list2_1, self.count_vac_by_year.values(), width_12, label='Количество вакансий')
            self.ax2.bar(x_list2_2, self.count_vac_by_year_for_prof.values(), width_12,
                         label=f'Количество вакансий\n{self.prof}')
            self.ax2.set_xticks(x_nums_2, self.count_vac_by_year.keys(), rotation='vertical')
            self.ax2.tick_params(axis='both', labelsize=8)
            self.ax2.legend(fontsize=8)
            self.ax2.grid(True, axis='y')

            list_names = {}
            for key, value in self.salary_lvl_by_city.items():
                if ' ' in key:
                    key = str(key).replace(' ', '\n')
                elif '-' in key and key.count('-') == 1:
                    key = str(key).replace('-', '-\n')
                elif '-' in key and key.count('-') != 1:
                    key = str(key).replace('-', '-\n', 1)
                list_names[key] = value

            width_3 = 0.7
            y_nums = np.arange(len(list(list_names.keys())))

            self.ax3.barh(y_nums, list_names.values(), width_3, align='center')
            self.ax3.set_yticks(y_nums, list_names.keys())
            self.ax3.tick_params(axis='y', labelsize=6)
            self.ax3.tick_params(axis='x', labelsize=8)
            self.ax3.invert_yaxis()
            self.ax3.grid(True, axis='x')

            other = 1
            data = [1]
            labels = ['Другие']
            for key, value in self.vacancy_rate_by_city.items():
                data.append(value * 100)
                labels.append(key)
                other -= value
            data[0] = round(other, 4) * 100
            textprops = {"fontsize": 6}

            self.ax4.pie(data, labels=labels, textprops=textprops, radius=1.1)

            plt.tight_layout()
            plt.savefig('graph.png')

        def generate_pdf(self):
            """Осуществляет генерацию pdf файла, содержащего графики из generate_image и данные в таблице из generate_excel.

            Returns:
                : pdf файл
            """
            env = Environment(loader=FileSystemLoader('.'))
            template = env.get_template('pdf_template.html')
            names_sheet1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {self.prof}',
                            'Количество вакансий', f'Количество вакансий - {self.prof}']
            names_sheet2 = ['Город', 'Уровень зарплат', 'Город', 'Доля вакансий']
            statistics_by_city_dic = []
            for i in range(len(list(self.salary_lvl_by_city.keys()))):
                statistics_by_city_dic.append(
                    [list(self.salary_lvl_by_city.keys())[i], list(self.vacancy_rate_by_city.keys())[i]])
            for key, value in self.vacancy_rate_by_city.items():
                self.vacancy_rate_by_city[key] = str(round(value * 100, 2)) + '%'
            pdf_template = template.render({'name': self.prof,
                                            'salary_lvl_by_year': self.salary_lvl_by_year,
                                            'count_vac_by_year': self.count_vac_by_year,
                                            'salary_lvl_by_year_for_prof': self.salary_lvl_by_year_for_prof,
                                            'count_vac_by_year_for_prof': self.count_vac_by_year_for_prof,
                                            'statistics_by_city_dic': statistics_by_city_dic,
                                            'salary_lvl_by_city': self.salary_lvl_by_city,
                                            'vacancy_rate_by_city': self.vacancy_rate_by_city,
                                            'names_sheet1': names_sheet1,
                                            'names_sheet2': names_sheet2})
            config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
            pdfkit.from_string(pdf_template, 'report.pdf', configuration=config,
                               options={"enable-local-file-access": ""})
            os.remove('graph.png')


    def sort_area_dict(dictionary):
        sorted_tuples = sorted(dictionary.items(), key=lambda item: item[1], reverse=True)[:10]
        sorted_dict = {k: v for k, v in sorted_tuples}
        return sorted_dict


    def output(df, years, vacancy):
        salary = Salary(df)
        df['salary'] = salary.df['salary']
        # df['salary'] = df.loc[:, ['salary_from', 'salary_to']].mean(axis=1)
        df['published_at'] = df['published_at'].apply(lambda x: int(x[:4]))
        # df_vacancy = df[df['name'].str.contains(vacancy)]

        total = len(df)
        df['count'] = df.groupby('area_name')['area_name'].transform('count')
        df_norm = df[df['count'] > 0.01 * total]
        cities = df_norm["area_name"].unique()

        salary_lvl_by_year, count_vac_by_year, salary_lvl_by_year_for_prof = {}, {}, {}
        count_vac_by_year_for_prof, salary_lvl_by_city, vacancy_rate_by_city = {}, {}, {}

        for city in cities:
            df_city = df_norm[df_norm['area_name'] == city]
            salary_lvl_by_city[city] = int(mean(df_city['salary']))
            vacancy_rate_by_city[city] = round(len(df_city) / len(df), 4)

        executor = futures.ProcessPoolExecutor()
        for year in years:
            statistics = executor.submit(analysis_statistics, vacancy, year).result()
            salary_lvl_by_year.update(statistics[0])
            count_vac_by_year.update(statistics[1])
            salary_lvl_by_year_for_prof.update(statistics[2])
            count_vac_by_year_for_prof.update(statistics[3])

        print("Динамика уровня зарплат по годам:", salary_lvl_by_year)
        print("Динамика количества вакансий по годам:", count_vac_by_year)
        print("Динамика уровня зарплат по годам для профессии:", salary_lvl_by_year_for_prof)
        print("Динамика количества вакансий по годам для профессии:", count_vac_by_year_for_prof)
        return [salary_lvl_by_year, count_vac_by_year, salary_lvl_by_year_for_prof, count_vac_by_year_for_prof, sort_area_dict(salary_lvl_by_city), sort_area_dict(vacancy_rate_by_city)]


    pd.set_option("expand_frame_repr", False)

    user_input = UserInput()
    csv_files = GetCsvParts(user_input.file_name)
    output_data = output(csv_files.df, csv_files.years, user_input.profession_name)
    report = Report(output_data, user_input.profession_name)
    report.generate_excel()
    report.generate_image()
    report.generate_pdf()
